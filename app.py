import streamlit as st
import pandas as pd
from copy import copy
from openpyxl import load_workbook
import io
import zipfile
from datetime import datetime
from pathlib import Path

st.set_page_config(
    page_title="Amazon FBA Manifest Generator",
    page_icon="📦",
    layout="centered",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

:root {
    --background: #f8fafc;
    --foreground: #0f172a;
    --card: #ffffff;
    --card-foreground: #0f172a;
    --muted: #f1f5f9;
    --muted-foreground: #64748b;
    --border: #e2e8f0;
    --input: #e2e8f0;
    --primary: #6366f1;
    --primary-foreground: #ffffff;
    --ring: #6366f1;
    --radius: 0.625rem;
}

*, *::before, *::after { box-sizing: border-box; }

html, body, .stApp {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    background: var(--background) !important;
    color: var(--foreground) !important;
}

#MainMenu, footer, header { visibility: hidden; }
[data-testid="stToolbar"] { display: none; }

.block-container {
    max-width: 640px !important;
    padding: 48px 24px 64px !important;
}

/* ── Typography & layout ── */
.page-header { margin-bottom: 28px; }

.badge {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    background: #eef2ff;
    border: 1px solid #c7d2fe;
    color: #4338ca;
    font-size: 0.75rem;
    font-weight: 500;
    border-radius: 9999px;
    padding: 4px 12px;
    margin-bottom: 16px;
}
.badge-dot {
    width: 6px;
    height: 6px;
    background: var(--primary);
    border-radius: 50%;
    flex-shrink: 0;
}

.page-title {
    margin: 0 0 8px;
    font-size: 1.875rem;
    font-weight: 700;
    letter-spacing: -0.025em;
    color: var(--foreground);
    line-height: 1.2;
}

.page-subtitle {
    margin: 0;
    font-size: 0.9375rem;
    color: var(--muted-foreground);
    line-height: 1.5;
    font-weight: 400;
}

/* ── Cards ── */
.ui-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 20px 24px;
    margin-bottom: 16px;
}

.ui-card-title {
    font-size: 0.9375rem;
    font-weight: 600;
    color: var(--card-foreground);
    margin: 0 0 4px;
}

.ui-card-desc {
    font-size: 0.8125rem;
    color: var(--muted-foreground);
    margin: 0 0 16px;
    line-height: 1.5;
}

.ui-card-desc:last-child { margin-bottom: 0; }

.section-label {
    font-size: 0.6875rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: var(--muted-foreground);
    margin: 0 0 12px;
}

/* ── Empty / preview state ── */
.state-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 48px 24px;
    text-align: center;
    margin-bottom: 16px;
}

.state-icon {
    width: 40px;
    height: 40px;
    margin: 0 auto 16px;
    color: #cbd5e1;
}

.state-title {
    font-size: 0.9375rem;
    font-weight: 600;
    color: var(--foreground);
    margin: 0 0 6px;
}

.state-desc {
    font-size: 0.8125rem;
    color: var(--muted-foreground);
    margin: 0;
    line-height: 1.5;
}

/* ── File meta ── */
.meta-row {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-bottom: 16px;
}

.meta-chip {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: var(--muted);
    border: 1px solid var(--border);
    border-radius: calc(var(--radius) - 2px);
    padding: 4px 10px;
    font-size: 0.75rem;
    font-weight: 500;
    color: var(--foreground);
}

.meta-chip .dot {
    width: 6px;
    height: 6px;
    background: #22c55e;
    border-radius: 50%;
}

/* ── Pack group grid ── */
.pg-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(72px, 1fr));
    gap: 8px;
}

.pg-tile {
    background: var(--muted);
    border: 1px solid var(--border);
    border-radius: calc(var(--radius) - 2px);
    padding: 12px 8px;
    text-align: center;
}

.pg-tile .pg-label {
    font-size: 0.625rem;
    font-weight: 600;
    color: var(--muted-foreground);
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-bottom: 4px;
}

.pg-tile .pg-count {
    font-size: 1.25rem;
    font-weight: 700;
    color: var(--foreground);
    line-height: 1;
}

.pg-tile .pg-unit {
    font-size: 0.625rem;
    color: var(--muted-foreground);
    margin-top: 2px;
}

/* ── File list ── */
.file-list {
    border: 1px solid var(--border);
    border-radius: calc(var(--radius) - 2px);
    overflow: hidden;
}

.file-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 10px 14px;
    font-size: 0.8125rem;
    border-bottom: 1px solid var(--border);
    background: var(--card);
}

.file-row:last-child { border-bottom: none; }

.file-row .fname {
    font-weight: 500;
    color: var(--foreground);
}

.file-row .fmeta {
    color: var(--muted-foreground);
    font-size: 0.75rem;
}

/* ── Date display ── */
.date-display {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    margin-top: 12px;
    padding: 8px 14px;
    background: var(--muted);
    border: 1px solid var(--border);
    border-radius: calc(var(--radius) - 2px);
    font-size: 0.875rem;
    font-weight: 500;
    color: var(--foreground);
}

/* ── Footer ── */
.page-footer {
    margin-top: 32px;
    padding-top: 20px;
    border-top: 1px solid var(--border);
    text-align: center;
    font-size: 0.75rem;
    color: var(--muted-foreground);
}

/* ── Streamlit overrides ── */
[data-testid="stFileUploader"] {
    margin-top: 0 !important;
}
[data-testid="stFileUploader"] section {
    border: 1px dashed var(--input) !important;
    border-radius: var(--radius) !important;
    background: #f8fafc !important;
    padding: 16px !important;
}
[data-testid="stFileUploader"] section:hover {
    border-color: #a5b4fc !important;
    background: #eef2ff !important;
}
[data-testid="stFileUploader"] section > div {
    padding: 0 !important;
}
[data-testid="stFileUploader"] button {
    background: var(--primary) !important;
    color: var(--primary-foreground) !important;
    border: none !important;
    border-radius: calc(var(--radius) - 2px) !important;
    font-size: 0.8125rem !important;
    font-weight: 500 !important;
    padding: 8px 16px !important;
    box-shadow: none !important;
}
[data-testid="stFileUploader"] button:hover {
    background: #4f46e5 !important;
    opacity: 1 !important;
}
[data-testid="stFileUploader"] small {
    color: var(--muted-foreground) !important;
    font-size: 0.75rem !important;
}

[data-testid="stDateInput"] label { display: none; }
[data-testid="stDateInput"] input {
    border-radius: calc(var(--radius) - 2px) !important;
    border: 1px solid var(--input) !important;
    font-size: 0.875rem !important;
    font-weight: 500 !important;
    padding: 8px 12px !important;
    background: var(--card) !important;
    color: var(--foreground) !important;
}
[data-testid="stDateInput"] input:focus {
    border-color: var(--ring) !important;
    box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.15) !important;
}

div[data-testid="stDownloadButton"] { margin: 0 !important; }
div[data-testid="stDownloadButton"] > button {
    width: 100% !important;
    background: var(--primary) !important;
    color: var(--primary-foreground) !important;
    border: none !important;
    border-radius: calc(var(--radius) - 2px) !important;
    font-size: 0.875rem !important;
    font-weight: 500 !important;
    padding: 10px 16px !important;
    box-shadow: none !important;
    transition: background 0.15s;
}
div[data-testid="stDownloadButton"] > button:hover {
    background: #4f46e5 !important;
    opacity: 1 !important;
}
div[data-testid="stDownloadButton"] > button[kind="secondary"] {
    background: var(--card) !important;
    color: var(--foreground) !important;
    border: 1px solid var(--input) !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important;
}
div[data-testid="stDownloadButton"] > button[kind="secondary"]:hover {
    background: var(--muted) !important;
}

[data-testid="stExpander"] {
    background: var(--card) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--radius) !important;
    overflow: hidden;
}
[data-testid="stExpander"] summary {
    font-size: 0.875rem !important;
    font-weight: 500 !important;
    color: var(--foreground) !important;
    padding: 12px 16px !important;
}
[data-testid="stExpander"] summary:hover {
    color: var(--primary) !important;
}
[data-testid="stExpander"] > div {
    border-top: 1px solid var(--border) !important;
    padding: 0 16px 16px !important;
}
[data-testid="stExpander"] p, [data-testid="stExpander"] li {
    font-size: 0.8125rem !important;
    color: var(--muted-foreground) !important;
    line-height: 1.6 !important;
}

[data-testid="stAlert"] {
    border-radius: var(--radius) !important;
    border: 1px solid #fecaca !important;
    background: #fef2f2 !important;
    font-size: 0.875rem !important;
}

.st-emotion-cache-1wmy9hl, .st-emotion-cache-16txtl3 {
    padding-top: 0 !important;
}
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"
RAW_TEMPLATE_PATH = Path(__file__).parent / "Amazon Uploader (Micro Tools).xlsx"
RAW_TEMPLATE_FILENAME = "Amazon Uploader (Micro Tools).xlsx"

DOC_ICON = """
<svg class="state-icon" xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.5">
  <path stroke-linecap="round" stroke-linejoin="round" d="M19.5 14.25v-2.625a3.375 3.375 0 0 0-3.375-3.375h-1.5A1.125 1.125 0 0 1 13.5 7.125v-1.5a3.375 3.375 0 0 0-3.375-3.375H8.25m2.25 0H5.625c-.621 0-1.125.504-1.125 1.125v17.25c0 .621.504 1.125 1.125 1.125h12.75c.621 0 1.125-.504 1.125-1.125V11.25a9 9 0 0 0-9-9Z"/>
</svg>
"""

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_raw(file) -> pd.DataFrame:
    df = pd.read_excel(file)
    df.columns = df.columns.str.strip()
    df["Pack Group #"] = df["Pack Group #"].ffill().astype(int)
    return df


def find_template_sheet(wb) -> str:
    for name in wb.sheetnames:
        n = name.lower()
        if "template" in n and "example" not in n:
            return name
    raise ValueError("Template sheet not found.")


def find_header_row(ws) -> int:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == "Merchant SKU":
                return cell.row
    return 6


def format_upc_code(value) -> str:
    text = str(value).strip()
    if "-FNSKU" in text:
        return text
    return str(int(value))


def aggregate_pg_rows(pg_rows: pd.DataFrame) -> pd.DataFrame:
    work = pg_rows.copy()
    work["UPC Code"] = work["UPC Code"].map(format_upc_code)
    work["_order"] = range(len(work))
    aggregated = (
        work.groupby("UPC Code", as_index=False)
        .agg(Total_QTY=("Total QTY", "sum"), _order=("_order", "min"))
        .rename(columns={"Total_QTY": "Total QTY"})
        .sort_values("_order")
        .drop(columns="_order")
    )
    aggregated["Total QTY"] = aggregated["Total QTY"].astype(int)
    return aggregated.reset_index(drop=True)


def build_manifest(template_bytes: bytes, pg_rows: pd.DataFrame) -> bytes:
    pg_rows = aggregate_pg_rows(pg_rows)
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb[find_template_sheet(wb)]

    # Ensure default prep/label owners are present on the template sheet
    # so that generated manifests always match the latest Amazon format.
    if ws.cell(row=3, column=1).value is None:
        ws.cell(row=3, column=1).value = "Default prep owner"
        ws.cell(row=3, column=2).value = "Seller"
    if ws.cell(row=4, column=1).value is None:
        ws.cell(row=4, column=1).value = "Default labeling owner"
        ws.cell(row=4, column=2).value = "Seller"

    hr = find_header_row(ws)
    style_ref_sku = ws.cell(row=hr + 1, column=1)
    style_ref_qty = ws.cell(row=hr + 1, column=2)
    sku_font = copy(style_ref_sku.font)
    sku_alignment = copy(style_ref_sku.alignment)
    qty_font = copy(style_ref_qty.font)
    qty_alignment = copy(style_ref_qty.alignment)
    qty_number_format = style_ref_qty.number_format

    for r in range(hr + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    for i, (_, row) in enumerate(pg_rows.iterrows()):
        r = hr + 1 + i
        sku_cell = ws.cell(row=r, column=1)
        qty_cell = ws.cell(row=r, column=2)
        sku_cell.value = row["UPC Code"]
        sku_cell.number_format = "@"
        sku_cell.font = copy(sku_font)
        sku_cell.alignment = copy(sku_alignment)
        qty_cell.value = int(row["Total QTY"])
        qty_cell.number_format = qty_number_format
        qty_cell.font = copy(qty_font)
        qty_cell.alignment = copy(qty_alignment)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Page header ───────────────────────────────────────────────────────────────
st.markdown("""
<div class="page-header">
    <div class="badge"><span class="badge-dot"></span>Amazon FBA · Send to Amazon</div>
    <h1 class="page-title">Manifest Generator</h1>
    <p class="page-subtitle">Upload your RAW Excel file, pick a shipment date, and download FBA manifests for every pack group.</p>
</div>
""", unsafe_allow_html=True)

# ── Upload card ───────────────────────────────────────────────────────────────
st.markdown("""
<div class="ui-card">
    <p class="ui-card-title">Upload RAW file</p>
    <p class="ui-card-desc">Select the filled Excel template from your shipment workflow, or download a blank template first.</p>
</div>
""", unsafe_allow_html=True)

st.download_button(
    label="Download blank template",
    data=RAW_TEMPLATE_PATH.read_bytes(),
    file_name=RAW_TEMPLATE_FILENAME,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="secondary",
    use_container_width=True,
)

st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

raw_file = st.file_uploader(
    "Upload RAW file",
    type=["xlsx", "xls"],
    label_visibility="collapsed",
)

# ── Date card ─────────────────────────────────────────────────────────────────
st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
st.markdown("""
<div class="ui-card">
    <p class="ui-card-title">Shipment date</p>
    <p class="ui-card-desc">Used in output filenames, e.g. <code style="background:#f1f5f9;padding:2px 6px;border-radius:4px;font-size:0.75rem;">CBN PG2 6.4.26.xlsx</code></p>
</div>
""", unsafe_allow_html=True)

selected_date = st.date_input("Date", value=datetime.today(), label_visibility="collapsed")
date_str = f"{selected_date.month}.{selected_date.day}.{str(selected_date.year)[-2:]}"
st.markdown(f'<div class="date-display">Shipment date · {date_str}</div>', unsafe_allow_html=True)

# ── Preview / results ─────────────────────────────────────────────────────────
st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

if raw_file:
    try:
        raw_file.seek(0)
        df = parse_raw(raw_file)
        vendor = str(df["Vendor"].iloc[0]).strip()
        pack_groups = sorted(df["Pack Group #"].unique())
        template_bytes = TEMPLATE_PATH.read_bytes()

        st.markdown(f"""
        <div class="ui-card">
            <p class="section-label">Preview</p>
            <div class="meta-row">
                <span class="meta-chip"><span class="dot"></span>{raw_file.name}</span>
                <span class="meta-chip">{vendor}</span>
                <span class="meta-chip">{len(pack_groups)} pack groups</span>
                <span class="meta-chip">{len(df)} SKUs</span>
            </div>
            <p class="ui-card-title" style="margin-bottom:12px;">Pack groups detected</p>
        </div>
        """, unsafe_allow_html=True)

        tiles_html = '<div class="pg-grid">'
        for pg in pack_groups:
            n_skus = len(df[df["Pack Group #"] == pg])
            tiles_html += f"""
            <div class="pg-tile">
                <div class="pg-label">PG {pg}</div>
                <div class="pg-count">{n_skus}</div>
                <div class="pg-unit">SKUs</div>
            </div>"""
        tiles_html += "</div>"
        st.markdown(tiles_html, unsafe_allow_html=True)

        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for pg in pack_groups:
                pg_rows = df[df["Pack Group #"] == pg][["UPC Code", "Total QTY"]]
                zf.writestr(
                    f"{vendor} PG{pg} {date_str}.xlsx",
                    build_manifest(template_bytes, pg_rows),
                )
        zip_buf.seek(0)

        st.markdown("""
        <div class="ui-card" style="padding-bottom:16px;">
            <p class="ui-card-title">Download manifests</p>
            <p class="ui-card-desc" style="margin-bottom:12px;">One Excel file per pack group, bundled in a ZIP archive.</p>
        </div>
        """, unsafe_allow_html=True)

        st.download_button(
            label=f"Download all {len(pack_groups)} manifests ({vendor} · {date_str})",
            data=zip_buf.getvalue(),
            file_name=f"{vendor} FBA Manifests {date_str}.zip",
            mime="application/zip",
            use_container_width=True,
        )

        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
        st.markdown('<p class="section-label">Files in ZIP</p>', unsafe_allow_html=True)

        rows_html = '<div class="file-list">'
        for pg in pack_groups:
            n_skus = len(df[df["Pack Group #"] == pg])
            fname = f"{vendor} PG{pg} {date_str}.xlsx"
            rows_html += f"""
            <div class="file-row">
                <span class="fname">{fname}</span>
                <span class="fmeta">{n_skus} SKUs</span>
            </div>"""
        rows_html += "</div>"
        st.markdown(rows_html, unsafe_allow_html=True)

    except Exception as exc:
        st.error(f"Error: {exc}")
        st.exception(exc)

else:
    st.markdown(f"""
    <div class="state-card">
        {DOC_ICON}
        <p class="state-title">No file uploaded</p>
        <p class="state-desc">Drop a RAW Excel file above to preview pack groups and generate manifests.</p>
    </div>
    """, unsafe_allow_html=True)

# ── How it works ──────────────────────────────────────────────────────────────
st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

with st.expander("How it works"):
    st.markdown("""
1. **Download the blank template** and fill in your SKU data — UPC Code, Vendor, quantities, and Pack Group #.
2. **Upload the completed RAW file** (.xlsx) using the upload area above.
3. **Set the shipment date** — this is used in every output filename.
4. **Download the ZIP** — you'll get one Amazon FBA manifest per pack group, ready for Seller Central.
    """)

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown(
    '<div class="page-footer">Manifest Generator · Maps UPC &amp; Quantity to Amazon upload templates</div>',
    unsafe_allow_html=True,
)
